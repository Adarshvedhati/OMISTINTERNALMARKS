# import os
# from openpyxl import Workbook, load_workbook


# def save_marks_to_excel(roll_no, name, course, subject, internal, marks):

#     file_name = f"{course}inter marks.xlsx"

#     # Create file if not exists
#     if not os.path.exists(file_name):

#         workbook = Workbook()
#         sheet = workbook.active

#         sheet.title = "Internal Marks"

#         sheet.append([
#             "Roll No",
#             "Name",
#             "Course",
#             "Subject",
#             "Internal",
#             "Marks"
#         ])

#         workbook.save(file_name)


#     # Open existing file
#     workbook = load_workbook(file_name)

#     sheet = workbook.active


#     # Add new student row
#     sheet.append([
#         roll_no,
#         name,
#         course,
#         subject,
#         internal,
#         marks
#     ])


#     workbook.save(file_name)
from openpyxl import Workbook

# ── Internal definitions per course ─────────────────────────────────────────
MCA_INTERNALS = ["Internal 1", "Internal 2"]
MBA_INTERNALS = ["Internal 1", "Internal 2", "Internal 3"]

INTERNALS_BY_COURSE = {
    "MCA": MCA_INTERNALS,
    "MBA": MBA_INTERNALS,
}


# ── Shared helpers ───────────────────────────────────────────────────────────

def get_college_filter(college_code: str, course: str) -> str:
    """2174 campus MBA students are stored under college_code 2177."""
    return "2177" if (college_code == "2174" and course == "MBA") else college_code


def build_marks_map(marks_qs) -> dict:
    """
    Collapse a Marks queryset into a nested lookup.

    Returns
    -------
    {student_id: {"Internal 1": <Marks obj>, "Internal 2": <Marks obj>, ...}}
    """
    marks_map: dict = {}
    for m in marks_qs:
        marks_map.setdefault(m.student_id, {})[m.internal] = m
    return marks_map


# ── openpyxl-based file builder ──────────────────────────────────────────────

def save_marks_to_excel(students, marks_map: dict, course: str,
                        subject: str, semester) -> str:
    """
    Build a combined internal-marks Excel file and save it to disk.

    Parameters
    ----------
    students  : queryset / list of Student objects, ordered by roll_no
    marks_map : output of build_marks_map()
    course    : "MCA" or "MBA"
    subject   : subject name  (used in filename)
    semester  : semester number (used in filename)

    Returns
    -------
    str : path to the saved .xlsx file

    MCA logic
    ---------
    - 2 internals
    - Columns: Roll No | Name |
               Int-1 Descriptive | Int-1 Assignment | Int-1 Total |
               Int-2 Descriptive | Int-2 Assignment | Int-2 Total |
               Average
    - Missing marks → 0; average = (Int-1 Total + Int-2 Total) / 2

    MBA logic
    ---------
    - 3 internals
    - Columns: Roll No | Name |
               Internal 1 Marks | Internal 2 Marks | Internal 3 Marks |
               Sum
    - Missing marks → 0; sum = Int-1 + Int-2 + Int-3
    """
    file_name = f"{course}_{subject}_sem{semester}_all_internals.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Internal Marks"

    if course == "MCA":
        ws.append([
            "Roll No", "Name",
            "Int-1 Descriptive (10-20)", "Int-1 Assignment (5-10)", "Int-1 Total",
            "Int-2 Descriptive (10-20)", "Int-2 Assignment (5-10)", "Int-2 Total",
            "Average",
        ])
        for stu in students:
            stu_marks = marks_map.get(stu.id, {})
            m1 = stu_marks.get("Internal 1")
            m2 = stu_marks.get("Internal 2")

            i1_desc  = m1.exam_marks       if m1 else 0
            i1_asn   = m1.assignment_marks if m1 else 0
            i1_total = m1.marks            if m1 else 0

            i2_desc  = m2.exam_marks       if m2 else 0
            i2_asn   = m2.assignment_marks if m2 else 0
            i2_total = m2.marks            if m2 else 0

            average = round((i1_total + i2_total) / 2, 2)

            ws.append([
                stu.roll_no, stu.name,
                i1_desc, i1_asn, i1_total,
                i2_desc, i2_asn, i2_total,
                average,
            ])

    else:  # MBA — 3 internals, show sum
        ws.append([
            "Roll No", "Name",
            "Internal 1 Marks", "Internal 2 Marks", "Internal 3 Marks",
            "Sum",
        ])
        for stu in students:
            stu_marks = marks_map.get(stu.id, {})
            m1 = stu_marks.get("Internal 1")
            m2 = stu_marks.get("Internal 2")
            m3 = stu_marks.get("Internal 3")

            i1 = m1.marks if m1 else 0
            i2 = m2.marks if m2 else 0
            i3 = m3.marks if m3 else 0

            ws.append([
                stu.roll_no, stu.name,
                i1, i2, i3,
                round(i1 + i2 + i3, 2),
            ])

    wb.save(file_name)
    return file_name